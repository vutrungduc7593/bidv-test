"""
Script to combine API test results from JSON and PNG files into an Excel file.
Creates two sheets:
1. Sheet 1: API Name and JSON Result
2. Sheet 2: API Name and Screenshot
"""

import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def extract_api_name(filename):
    """Extract API name from filename like 'api_result_API_ 100.json'"""
    match = re.search(r'api_result_(API_\s*\d+)', filename)
    if match:
        return match.group(1)
    return filename

def get_sorted_files(directory, extension):
    """Get sorted list of files with given extension"""
    files = list(directory.glob(f'*{extension}'))
    # Sort by the API number
    def sort_key(filepath):
        match = re.search(r'API_\s*(\d+)', filepath.name)
        if match:
            return int(match.group(1))
        return 0
    return sorted(files, key=sort_key)

def main():
    # Set up paths
    data_dir = Path('./data/api-test-result')
    output_file = Path('./data/api_test_results.xlsx')
    
    if not data_dir.exists():
        print(f"Error: Directory {data_dir} does not exist")
        return
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Sheet 1: JSON Results
    print("Creating Sheet 1: JSON Results...")
    ws1 = wb.create_sheet('JSON Results', 0)
    
    # Set headers for Sheet 1
    ws1['A1'] = 'API Name'
    ws1['B1'] = 'API Result'
    ws1['A1'].font = Font(bold=True, size=12)
    ws1['B1'].font = Font(bold=True, size=12)
    
    # Set column widths
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 100
    
    # Read and add JSON files
    json_files = get_sorted_files(data_dir, '.json')
    row = 2
    
    for json_file in json_files:
        api_name = extract_api_name(json_file.name)
        
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                json_content = json.load(f)
                # Format JSON with indentation for better readability
                json_str = json.dumps(json_content, indent=2, ensure_ascii=False)
            
            ws1[f'A{row}'] = api_name
            ws1[f'B{row}'] = json_str
            ws1[f'A{row}'].alignment = Alignment(vertical='top')
            ws1[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)
            
            print(f"  Added {api_name}")
            row += 1
            
        except Exception as e:
            print(f"  Error processing {json_file.name}: {e}")
    
    # Create Sheet 2: Screenshots
    print("\nCreating Sheet 2: Screenshots...")
    ws2 = wb.create_sheet('Screenshots', 1)
    
    # Set headers for Sheet 2
    ws2['A1'] = 'API Name'
    ws2['B1'] = "API Result's Screenshot"
    ws2['A1'].font = Font(bold=True, size=12)
    ws2['B1'].font = Font(bold=True, size=12)
    
    # Set column widths
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 80
    
    # Read and add PNG files
    png_files = get_sorted_files(data_dir, '.png')
    row = 2
    
    for png_file in png_files:
        api_name = extract_api_name(png_file.name)
        
        try:
            # Add API name
            ws2[f'A{row}'] = api_name
            ws2[f'A{row}'].alignment = Alignment(vertical='top')
            
            # Add image
            img = Image(str(png_file))
            
            # Resize image to fit in cell (max width 600 pixels, maintain aspect ratio)
            max_width = 600
            if img.width > max_width:
                ratio = max_width / img.width
                img.width = max_width
                img.height = int(img.height * ratio)
            
            # Set row height to accommodate image (height in points, 1 point ≈ 1.33 pixels)
            row_height = (img.height / 1.33) + 10
            ws2.row_dimensions[row].height = row_height
            
            # Anchor image to cell B{row}
            img.anchor = f'B{row}'
            ws2.add_image(img)
            
            print(f"  Added {api_name}")
            row += 1
            
        except Exception as e:
            print(f"  Error processing {png_file.name}: {e}")
    
    # Save workbook
    print(f"\nSaving workbook to {output_file}...")
    wb.save(output_file)
    print(f"✓ Successfully created {output_file}")
    print(f"  - Sheet 1: {len(json_files)} JSON results")
    print(f"  - Sheet 2: {len(png_files)} screenshots")

if __name__ == '__main__':
    main()

